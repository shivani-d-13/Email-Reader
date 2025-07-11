import os
import base64
import re
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from dotenv import load_dotenv
import openpyxl
import json
from datetime import datetime

load_dotenv()
EMAIL_ADDRESS = os.getenv('EMAIL_ADDRESS')

SCOPES = ['https://www.googleapis.com/auth/gmail.modify']
CREDENTIALS_FILE = 'credentials.json'
TOKEN_FILE = 'token.json'

def authenticate():
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, 'w') as token:
            token.write(creds.to_json())
    return creds

def search_emails(service, subject):
    query = f'subject:"{subject}" is:unread'
    result = service.users().messages().list(userId='me', q=query).execute()
    messages = result.get('messages', [])
    return messages

def get_excel_attachment(service, message_id):
    message = service.users().messages().get(userId='me', id=message_id).execute()
    for part in message['payload'].get('parts', []):
        if 'filename' in part and part['filename'].endswith('.xlsx'):
            attachment_id = part['body']['attachmentId']
            attachment = service.users().messages().attachments().get(
                userId='me', messageId=message_id, id=attachment_id).execute()
            data = base64.urlsafe_b64decode(attachment['data'])
            with open('attachment.xlsx', 'wb') as f:
                f.write(data)
            print(f"Attachment saved as 'attachment.xlsx'")
            return 'attachment.xlsx'
    print("No Excel attachment found.")
    return None

def mark_as_read(service, message_id):
    service.users().messages().modify(
        userId='me',
        id=message_id,
        body={'removeLabelIds': ['UNREAD']}
    ).execute()
    print("Marked email as read.")


def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    data = []
    headers = [cell.value for cell in next(sheet.iter_rows(values_only=True))]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        data.append(row_dict)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_filename = f'output_{timestamp}.json'

    with open(json_filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

    print(f"Excel data saved to '{json_filename}'")


def main():
    creds = authenticate()
    service = build('gmail', 'v1', credentials=creds)
    subject_to_search = "Job Register"  # Change to your subject
    messages = search_emails(service, subject_to_search)
    if not messages:
        print("No unread emails found with that subject.")
        return

    latest_message_id = messages[0]['id']
    file_path = get_excel_attachment(service, latest_message_id)
    if file_path:
        read_excel(file_path)
        mark_as_read(service, latest_message_id)

if __name__ == '__main__':
    main()
